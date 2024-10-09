import openpyxl
import pandas as pd
import numpy as np
from decorators import time_check
from link import update_aewiki_links, update_seesaa_links
from config import (
    FILE_NAME,
    CHARACTER_SHEET_NAME,
    BUDDY_SHEET_NAME,
    PERSONALITY_SHEET_NAME,
    DUNGEON_SHEET_NAME,
    TRANSLATE_SHEET_NAME,
    RESULT_PATH,
    PARSE_REF_DICT,
    IMG_SOURCE_PATH,
    AC_MAPPING,
    SEESAA_MAPPING,
)
from function import init_folder, write_json
import sys
import os
import shutil
import requests
import urllib.parse
from openpyxl.workbook.workbook import Workbook
from pathlib import Path
from PIL import Image


def find_personality_tag(df: pd.DataFrame, word: str) -> str:
    filtered_df = df[df["jap"] == word]
    if not filtered_df.empty:
        return filtered_df.iloc[0, 0]
    else:
        raise Exception("personality tag not found")

def find_dungeon_id(df: pd.DataFrame, word: str) -> int:
    filtered_df = df[df["name"] == word]
    if not filtered_df.empty:
        return int(filtered_df.iloc[0, 0])
    else:
        raise Exception("dungeon id not found")

@time_check
def make_data(wb: Workbook):
    for sheet_nm in [
        CHARACTER_SHEET_NAME,
        BUDDY_SHEET_NAME,
        PERSONALITY_SHEET_NAME,
        DUNGEON_SHEET_NAME,
        TRANSLATE_SHEET_NAME,
    ]:
        if sheet_nm not in wb:
            print("시트가 없습니다.")
            sys.exit()

    cha_sheet = wb[CHARACTER_SHEET_NAME]
    bud_sheet = wb[BUDDY_SHEET_NAME]
    tsl_sheet = wb[TRANSLATE_SHEET_NAME]
    
    translate_df = pd.read_excel(FILE_NAME, sheet_name=TRANSLATE_SHEET_NAME, dtype=str, engine='openpyxl')
    dungeon_df = pd.read_excel(FILE_NAME, sheet_name=DUNGEON_SHEET_NAME, dtype=str, engine='openpyxl')
    personality_df = pd.read_excel(FILE_NAME, sheet_name=PERSONALITY_SHEET_NAME, dtype=str, engine='openpyxl').replace({np.nan: None})
    
    os.makedirs("code/links", exist_ok=True)
    update_seesaa_links()
    update_aewiki_links()

    with open("code/links/aewiki.txt", "r+", encoding="utf-8") as f:
        aewiki_links = list(set(map(str.strip, f.readlines())))
    with open("code/links/seesaa.txt", "r+", encoding="utf-8") as f:
        seesaa_links = list(set(map(str.strip, f.readlines())))

    except_chars = ["シオン (テイルズ)", "サイラス (オクトパス)"]

    """
    1. Character & Quick Search
    """
    print("\nadd characters")
    char_json = []
    char_names = []
    for row in cha_sheet.iter_rows(min_row=2):
        if row[0].value is None: break
        id: int             = int(row[0].value)
        style: str          = str(row[2].value)
        code: int           = int(row[3].value)
        get: str            = str(row[4].value).upper()
        type: str           = str(row[5].value)
        from_arr: list[int]     = list(map(int, str(row[6].value).split(","))) if row[6].value else []
        change_arr: list[int]   = list(map(int, str(row[7].value).split(","))) if row[7].value else []
        drop_arr: list[str]     = list(map(lambda x: find_dungeon_id(dungeon_df, x), row[9].value.split(",") if row[9].value else []))
        alter: str          = "TRUE" if row[10].value else "FALSE"
        manifest: str       = row[11].value or "없음"
        staralign: str          = "TRUE" if row[12].value else "FALSE"

        tags: list[str] = [
            PARSE_REF_DICT["style"][style],
            PARSE_REF_DICT["get"][get],
            PARSE_REF_DICT["type"][type],
            PARSE_REF_DICT["alter"][alter],
            PARSE_REF_DICT["manifest"][manifest],
            PARSE_REF_DICT["staralign"][staralign],
        ]

        keyword: str = ""
        eng_keyword: str = ""
        for row in tsl_sheet.iter_rows(min_row=2):
            if row[0].value == f"c{code}":
                keyword = row[2].value.split(" ")[0] if row[2].value not in except_chars else row[2].value
                eng_keyword = row[3].value.split(" ")[0] if "(Tales)" in row[3].value else row[3].value
                if keyword in AC_MAPPING.keys():
                    eng_keyword = AC_MAPPING[keyword]
                if keyword in SEESAA_MAPPING.keys():
                    keyword = SEESAA_MAPPING[keyword]
                break
        if len(keyword) <= 0:
            print(f"번역 정보가 없습니다. id : {id}")
            sys.exit()
        if style in ["AS", "ES"]:
            keyword += f"({style})"
            eng_keyword += f"({style})"
        
        filtered_personality_df = personality_df.loc[personality_df['name'] == keyword]
        if filtered_personality_df.empty:
            raise ValueError(f"{keyword}의 퍼스널리티 정보가 없습니다.")
        else:
            p_tag_list, year = filtered_personality_df.iloc[0, [1,2]]
            tags += list(map(lambda x: find_personality_tag(translate_df, x), p_tag_list.split(",")))

        # 업데이트 지연에 따른 조건문...
        if id <= 392:
            # Seesaa Wiki Mapping
            seesaa_endpoint = keyword.replace("(AS)", "(アナザースタイル)") \
                .replace("(ES)", "(エクストラスタイル)") \
                .replace(" ", "")
            
            seesaa_parsed = urllib.parse.quote(seesaa_endpoint.encode("euc-jp"))
            try:
                seesaa_url = next(f'https://anothereden.game-info.wiki/d/{seesaa_parsed}' for link in seesaa_links if link == f'https://anothereden.game-info.wiki/d/{seesaa_parsed}'.lower())
            except StopIteration:
                seesaa_parsed = urllib.parse.quote(seesaa_endpoint.replace("(アナザースタイル)", "").encode("euc-jp"))
                seesaa_url = next(f'https://anothereden.game-info.wiki/d/{seesaa_parsed}' for link in seesaa_links if link == f'https://anothereden.game-info.wiki/d/{seesaa_parsed}'.lower())
            
            # AE Wiki Mapping
            aewiki_endpoint = eng_keyword.replace("(AS)", "_(Another Style)") \
                .replace("(ES)", "_(Extra Style)") \
                .replace(" ", "_")        
                
            aewiki_url = f'https://anothereden.wiki/w/{aewiki_endpoint}'
            if "(Alter)" in aewiki_endpoint or "Strawboy" in aewiki_endpoint:            
                test_res = requests.get(aewiki_url)
                if not test_res.ok:
                    print(f'\nAE Wiki url 오류! 이름 : {eng_keyword}')
                    sys.exit()
            else:
                aewiki_url = next(aewiki_url for link in aewiki_links if link == f'/w/{aewiki_endpoint}')
        else:
            seesaa_url, aewiki_url = None, None
        
        char_json.append({
            "id": id,
            "code": code,
            "tags": sorted(tags),
            "from": from_arr,
            "change": change_arr,
            "dungeon_drop": drop_arr,
            "year": year,
            "seesaa": seesaa_url,
            "aewiki": aewiki_url
        })

        char_names.append(keyword)
        print(".", end="", flush=True)     
    
    personality_df = personality_df[~personality_df["name"].isin(char_names)]

    # Add absent characters
    for idx, row_num in enumerate(personality_df.index):
        jap_name = personality_df.loc[row_num, "name"]
        tags = list(map(lambda x: find_personality_tag(translate_df, x), personality_df.loc[row_num, "personalities"].split(",")))

        char_translate_df = translate_df.loc[(translate_df['jap'] == jap_name) & (translate_df["tag"].str.startswith("c"))]
        if char_translate_df.empty:
            raise Exception(f"{jap_name}의 번역 정보가 없습니다.")
        else:
            code = int(char_translate_df.iloc[0, 0].replace("c", ""))
            char_json.append({
                "id": idx + 1000,
                "code": code,
                "tags": sorted(tags),
            })
            print(".", end="", flush=True)  

    write_json(f"{RESULT_PATH}/data/character.json", char_json)


    """
    2. Buddy
    """
    print("\nadd buddies")
    buddy_json = []
    for row in bud_sheet.iter_rows(min_row=2):
        if row[0].value is None: break
        id: int                 = int(row[0].value)
        code: int               = row[2].value
        link_arr: list[int]     = list(map(int, str(row[3].value).split(","))) if row[3].value else []
        get: str = "get.notfree" if len(link_arr) > 0 else f"get.buddy{id}"

        buddy_translate_df = translate_df[translate_df['tag'] == f"bud{code}"]
        if buddy_translate_df.empty:
            raise Exception(f"bud{code} not found in translate sheet")
        else:
            keyword, eng_keyword = buddy_translate_df.iloc[0, [2,3]]
        
        # Seesaa Wiki Mapping
        seesaa_parsed = urllib.parse.quote(keyword.encode("euc-jp"))
        seesaa_url = next(
            (f'https://anothereden.game-info.wiki/d/{seesaa_parsed}' for link in seesaa_links if link == f'https://anothereden.game-info.wiki/d/{seesaa_parsed}'.lower()),
            None
        )
            
        # AE Wiki Mapping
        aewiki_endpoint = eng_keyword.replace(" ", "_").replace("(AS)", "_(Another_Style)")
        if eng_keyword in ["Iridian", "Thysía"]: 
            aewiki_url = f'https://anothereden.wiki/w/{aewiki_endpoint}'
            test_res = requests.get(aewiki_url)
            if not test_res.ok:
                raise Exception(f'\nAE Wiki url 오류! 이름 : {eng_keyword}')
        else:
            aewiki_url = next(
                (f'https://anothereden.wiki/w/{aewiki_endpoint}' for link in aewiki_links if link == f'/w/{aewiki_endpoint}'),
                None
            )

        buddy_json.append({
            "id": id,
            "code": code,
            "link": link_arr,
            "get": get,
            "seesaa": seesaa_url,
            "aewiki": aewiki_url
        })    
        print(".", end="", flush=True)

    print()
    write_json(f"{RESULT_PATH}/data/buddy.json", buddy_json)    


@time_check
def make_dungeon_data():
    df = pd.read_excel(FILE_NAME, sheet_name=DUNGEON_SHEET_NAME, engine='openpyxl')
    df = df.reindex(["id", "wiki", "altema"], axis=1) \
         .replace({np.nan: None})
    write_json(f"{RESULT_PATH}/data/dungeon.json", df.to_dict('records'))

@time_check
def make_i18n_data():
    df = pd.read_excel(FILE_NAME, sheet_name=TRANSLATE_SHEET_NAME, engine='openpyxl', index_col=0)
    if df.isnull().sum().sum() != 1:
        # 무속성에 None 있어서 하나 카운트됨;;
        raise Exception("null value in translate sheet")
    i18n_data = df.replace({np.nan: "None"}).T.to_dict('records')
    for idx, name in enumerate(["ko", "jp", "en"]):
        write_json(f"{RESULT_PATH}/i18n/{name}.json", i18n_data[idx])

@time_check
def copy_character_image():
    df = pd.read_json(f"{RESULT_PATH}/data/character.json")
    df = df.iloc[:, :3]
    if df.isnull().values.any():
        raise Exception("null value in dataframe")
    for val in df.values:
        id, code, tags = val
        if "style.extra" in tags:
            style = "ES"
        elif "style.another" in tags:
            style = "AS"
        elif "style.normal" in tags:
            style = "NS"
        else:
            style = "4.5"
        filename = "{}{}.png".format(code, PARSE_REF_DICT["img_parse"][style])
        awaken_filename = "{}{}_opened.png".format(code, PARSE_REF_DICT["img_parse"][style])
        
        # 아이디대로 이미지를 복사
        for file in os.listdir(IMG_SOURCE_PATH):
            if filename == file:
                shutil.copyfile(f"rawimage/{file}", f"{RESULT_PATH}/image/{id}.png")
            elif awaken_filename == file:
                shutil.copyfile(f"rawimage/{file}", f"{RESULT_PATH}/image/{id}_awaken.png")

@time_check
def copy_buddy_image():
    df = pd.read_excel(FILE_NAME, sheet_name=BUDDY_SHEET_NAME, dtype=str, engine='openpyxl')
    df = df.iloc[:, [0,2]]
    if df.isnull().values.any():
        raise Exception("null value in dataframe")
    for val in df.values:
        id, code = val
        shutil.copyfile(f"{IMG_SOURCE_PATH}/{code}.png", f"{RESULT_PATH}/image/{id}.png")

@time_check
def optimize_image():
    image_dir = f"{RESULT_PATH}/image"
    for file in os.listdir(image_dir):
        file_path = Path(f"{image_dir}/{file}")
        image = Image.open(file_path)
        if "awaken" in file:
            image.crop((15, 15, 120, 120))
        else:
            image.thumbnail((120, 120))
        webp_path = file_path.with_suffix(".webp")
        image.save(webp_path)
        image.save(file_path)

if __name__ == "__main__":
    folder_paths = [
        f"{RESULT_PATH}/image",
        f"{RESULT_PATH}/data",
        f"{RESULT_PATH}/i18n",
    ]

    for path in folder_paths:
        init_folder(path=path)
    wb = openpyxl.open(FILE_NAME, data_only=True, read_only=True)

    make_data(wb=wb)
    make_dungeon_data()
    make_i18n_data()
    copy_character_image()
    copy_buddy_image()
    optimize_image()