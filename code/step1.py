import openpyxl
import pandas as pd
from openpyxl.workbook.workbook import Workbook
from config import (
    FILE_NAME,
    CHARACTER_SHEET_NAME,
    BUDDY_SHEET_NAME,
    PERSONALITY_SHEET_NAME,
    DUNGEON_SHEET_NAME,
    TRANSLATE_SHEET_NAME,
)
import sys



def add_personality_tags() -> None:
    personality_df = pd.read_excel(FILE_NAME, sheet_name=PERSONALITY_SHEET_NAME, dtype=str, engine='openpyxl')
    translate_df = pd.read_excel(FILE_NAME, sheet_name=TRANSLATE_SHEET_NAME, dtype=str, engine='openpyxl')
    
    if personality_df.iloc[:,1].isnull().values.any():
        raise Exception("null value in personality")
    
    all_p_string = ",".join(list(personality_df.iloc[:,1]))
    p_list = list(set(all_p_string.split(",")))
    
    per_translate_df = translate_df.loc[(translate_df["tag"].str.startswith("personality."))]
    tagged_p_list = list(set(per_translate_df.iloc[:,2]))

    new_p_list = set(p_list) - set(tagged_p_list)
    if new_p_list:
        print("New personality:", new_p_list)
        tagged_p_count = len(tagged_p_list)
        for idx, p in enumerate(new_p_list):
            translate_df = pd.concat([
                translate_df, 
                pd.DataFrame([{"tag": f"personality.{idx+1+tagged_p_count}", "jap": p}])
            ], ignore_index=True)

        with pd.ExcelWriter(FILE_NAME, mode="a") as writer:
            translate_df.to_excel(writer, sheet_name=TRANSLATE_SHEET_NAME, index=False)

def add_char_book_tags() -> None:
    character_df = pd.read_excel(FILE_NAME, sheet_name=CHARACTER_SHEET_NAME, dtype=str, engine='openpyxl')
    translate_df = pd.read_excel(FILE_NAME, sheet_name=TRANSLATE_SHEET_NAME, dtype=str, engine='openpyxl')

    char_names = set(character_df.iloc[:,1])
    translated_names = set(translate_df.loc[(translate_df["tag"].str.startswith("c10"))].iloc[:,1])

    new_names = char_names - translated_names
    if new_names:
        for name in new_names:
            print("New character:", new_names)
            char_rows: pd.DataFrame = character_df[character_df["name"] == name]
            translate_df = pd.concat([
                translate_df, 
                pd.DataFrame([{"tag": f"c{char_rows.iloc[0,3]}", "kor": name}])
            ], ignore_index=True)

        with pd.ExcelWriter(FILE_NAME, mode="a", if_sheet_exists="overlay") as writer:
            translate_df.to_excel(writer, sheet_name=TRANSLATE_SHEET_NAME, index=False)
    
    char_books = set(character_df.loc[character_df["book"].notnull()].iloc[:,8])
    translated_books = set(translate_df.loc[(translate_df["tag"].str.startswith("book.char"))].iloc[:,1])

    new_books = char_books - translated_books
    if new_books:
        for book in new_books:
            print("New book:", book)
            char_rows: pd.DataFrame = character_df[character_df["book"] == book]
            translate_df = pd.concat([
                translate_df, 
                pd.DataFrame([{"tag": f"book.char{char_rows.iloc[0,0]}", "kor": book}])
            ], ignore_index=True)

        with pd.ExcelWriter(FILE_NAME, mode="a", if_sheet_exists="overlay", engine="openpyxl", engine_kwargs={ "data_only": True }) as writer:
            translate_df.to_excel(writer, sheet_name=TRANSLATE_SHEET_NAME, index=False)


def add_buddy_tags(wb: Workbook) -> None:
    if BUDDY_SHEET_NAME not in wb or TRANSLATE_SHEET_NAME not in wb:
        print("참조할 시트가 없습니다.")
        sys.exit()
    
    b_sheet = wb[BUDDY_SHEET_NAME]
    t_sheet = wb[TRANSLATE_SHEET_NAME]

    buddy_dic = {}
    for row in b_sheet.iter_rows(min_row=2):
        name = row[1].value
        code = row[2].value
    
        buddy_dic[code] = name

    # 번역시트의 1열 = 태그 
    key_list = list(filter(lambda x: x is not None and len(x) > 0, [i[0].value for i in t_sheet.iter_rows(min_row=2)]))
    for code in buddy_dic:
        tag = "bud{}".format(code)
        if tag not in key_list:
            print("add tag", tag)
            kor_name = buddy_dic[code]
            t_sheet.append([tag, kor_name])



def add_dungeon_tags(wb: Workbook) -> None:
    if DUNGEON_SHEET_NAME not in wb or TRANSLATE_SHEET_NAME not in wb:
        print("참조할 시트가 없습니다.")
        sys.exit()
    
    d_sheet = wb[DUNGEON_SHEET_NAME]
    t_sheet = wb[TRANSLATE_SHEET_NAME]

    # 한국어명 dic 작성
    dungeon_dic = {}
    for row in d_sheet.iter_rows(min_row=2):
        id = row[0].value
        name = row[1].value
    
        dungeon_dic[id] = name
 
    # 시트의 1열 = key값 = 태그 
    key_list = list(filter(lambda x: x is not None and len(x) > 0, [i[0].value for i in t_sheet.iter_rows(min_row=2)]))
    for id in dungeon_dic:
        id_str = str(id).zfill(3)
        tag = "drop.dungeon{}".format(id_str)
        if tag not in key_list:
            print("add tag", tag)
            kor_name = dungeon_dic[id]
            t_sheet.append([tag, kor_name])

def make_translate_sheet() -> None:
    add_personality_tags()
    add_char_book_tags()

    wb = openpyxl.open(FILE_NAME, data_only=True)

    add_buddy_tags(wb=wb)
    add_dungeon_tags(wb=wb)

    wb.save(FILE_NAME)



# execute
if __name__ == "__main__":    
    make_translate_sheet()